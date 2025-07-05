import pandas as pd
from openpyxl import load_workbook
import shutil
import os

# File paths (use raw strings to avoid unicode escape errors)
erection_data_sample_path = r"C:\Users\twink\OneDrive\Desktop\ANJAR PROJECT\CODING\DIAGRAMS\For Code\61) FOR ERECTION FEEDERWISE BOQ\ANJAR FINAL.xlsx"
erection_format_path = r"C:\Users\twink\OneDrive\Desktop\ANJAR PROJECT\CODING\DIAGRAMS\For Code\61) FOR ERECTION FEEDERWISE BOQ\ERECTION FORMAT.xlsx"
output_dir = r"C:\Users\twink\OneDrive\Desktop\ANJAR PROJECT\CODING\DIAGRAMS\For Code\61) FOR ERECTION FEEDERWISE BOQ\ERECTION_SHEETS"

# Create the output directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

# Load the erection data sample workbook
data_wb = load_workbook(erection_data_sample_path, data_only=True)
sheet_names = data_wb.sheetnames

saved_files = []

# Loop through each sheet in the data sample file
for sheet_name in sheet_names:
    # Copy the template workbook for modification
    modified_wb_path = os.path.join(output_dir, f"{sheet_name}.xlsx")
    shutil.copy(erection_format_path, modified_wb_path)
    modified_wb = load_workbook(modified_wb_path)
    modified_ws = modified_wb.active
    
    data_ws = data_wb[sheet_name]
    
    # Perform the cell updates as per instruction
    modified_ws["A6"] = data_ws["A6"].value
    modified_ws["K10"] = data_ws["N9"].value
    modified_ws["K11"] = data_ws["N12"].value
    modified_ws["K12"] = data_ws["N13"].value
    modified_ws["K13"] = data_ws["N20"].value
    modified_ws["K14"] = data_ws["N21"].value
    modified_ws["K16"] = data_ws["N22"].value
    modified_ws["K17"] = data_ws["N23"].value
    
    # Save the modified workbook
    modified_wb.save(modified_wb_path)
    saved_files.append(modified_wb_path)

# Print confirmation for each saved file
for file in saved_files:
    print("Saved:", file)
